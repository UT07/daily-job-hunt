"""Master Excel Application Tracking System.

A single persistent Excel file that accumulates ALL jobs across every run.
Your central hub for tracking the entire application lifecycle:
- Job discovery with 3-score breakdown (ATS, HM, TR)
- Tailored resume + cover letter PDFs (linked to S3)
- LinkedIn contacts with connection messages + follow-up drafts
- Application status tracking (Applied, Interview, Offer, etc.)
- "Apply Reminder" column highlights unapplied matched jobs
- Follow-up date tracking (1 week and 2 weeks after applying)
- Daily Summary sheet with aggregate stats
- Deduplication: same job+company combo is never added twice
"""

from __future__ import annotations
import json
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from scrapers.base import Job

logger = logging.getLogger(__name__)


# ── Column definitions ──────────────────────────────────────────────────────
COLUMNS = [
    ("Date Found", 12),
    ("Posted Date", 12),
    ("Match", 8),
    ("Score", 8),
    ("ATS", 7),
    ("HM", 7),
    ("TR", 7),
    ("Title", 30),
    ("Company", 22),
    ("Location", 20),
    ("Remote?", 9),
    ("Salary", 18),
    ("Source", 10),
    ("Resume Type", 14),
    ("Apply Link", 15),
    ("Resume PDF", 20),
    ("Cover Letter", 20),
    ("Resume (Drive)", 20),
    ("CL (Drive)", 20),
    # Networking columns
    ("Contact 1", 25),
    ("Contact 1 LinkedIn", 15),
    ("Contact 1 Message", 40),
    ("Contact 2", 25),
    ("Contact 2 LinkedIn", 15),
    ("Contact 2 Message", 40),
    ("Contact 3", 25),
    ("Contact 3 LinkedIn", 15),
    # Application tracking
    ("Applied?", 10),
    ("Applied Date", 13),
    ("Status", 14),
    ("Follow-Up 1", 13),
    ("Follow-Up 2", 13),
    ("Follow-Up 1 Msg", 45),
    ("Follow-Up 2 Msg", 45),
    ("Apply Reminder", 15),
    ("Notes", 35),
]

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SCORE_EXCELLENT = PatternFill("solid", fgColor="92D050")  # Green: 85+
SCORE_GOOD = PatternFill("solid", fgColor="C6EFCE")       # Light green: 75-84
SCORE_OK = PatternFill("solid", fgColor="FFEB9C")          # Yellow: 60-74
SCORE_LOW = PatternFill("solid", fgColor="FFC7CE")         # Red: <60

REMINDER_FILL = PatternFill("solid", fgColor="FF6B6B")     # Red: "APPLY NOW!"
FOLLOWUP_DUE = PatternFill("solid", fgColor="FFD93D")      # Yellow: follow-up due

STATUS_COLORS = {
    "New": PatternFill("solid", fgColor="D9E2F3"),
    "Applied": PatternFill("solid", fgColor="B4C6E7"),
    "Interview": PatternFill("solid", fgColor="E2EFDA"),
    "Offer": PatternFill("solid", fgColor="92D050"),
    "Rejected": PatternFill("solid", fgColor="FFC7CE"),
    "Withdrawn": PatternFill("solid", fgColor="D9D9D9"),
}

BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")


def create_or_update_tracker(
    jobs: List[Job],
    tracker_path: str,
    run_date: str = None,
) -> str:
    """Create or append to the master job tracker Excel file.

    Deduplicates by title+company and updates apply reminders + follow-up dates.
    """
    tracker_path = Path(tracker_path)
    run_date = run_date or datetime.now().strftime("%Y-%m-%d")

    if tracker_path.exists():
        wb = load_workbook(str(tracker_path))
        ws = wb.active
        existing_keys = _get_existing_keys(ws)
        start_row = ws.max_row + 1
        # Update reminders and follow-ups for existing rows
        _update_reminders(ws, run_date)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Tracker"
        _setup_header(ws)
        _add_data_validations(ws)
        start_row = 2
        existing_keys = set()

        summary = wb.create_sheet("Daily Summary")
        _setup_summary_sheet(summary)

    # Filter out duplicates
    new_jobs = []
    for job in jobs:
        key = f"{job.title.lower().strip()}|{job.company.lower().strip()}"
        if key not in existing_keys:
            new_jobs.append(job)
            existing_keys.add(key)

    skipped = len(jobs) - len(new_jobs)
    if skipped > 0:
        logger.info(f"[EXCEL] Skipped {skipped} duplicate jobs already in tracker")

    # Add new job rows
    for i, job in enumerate(new_jobs):
        row = start_row + i
        is_zebra = (row % 2 == 0)

        # Parse contacts (up to 3)
        contacts = []
        if job.linkedin_contacts:
            try:
                contacts = json.loads(job.linkedin_contacts)
            except json.JSONDecodeError:
                pass

        # Follow-up message drafts
        followup_1_msg = f"Hi! I applied for the {job.title} role at {job.company} last week and wanted to follow up. I'm very excited about this opportunity and would love to discuss how my experience aligns. Would you have a few minutes for a quick chat?"
        followup_2_msg = f"Hi! I hope you're well. I wanted to circle back on my application for the {job.title} position at {job.company}. I remain very interested and confident I'd be a great fit. Happy to provide any additional information that might be helpful."

        # File naming for display
        safe_role = "".join(c for c in job.title if c.isalnum() or c in " _-")[:25].strip().replace(" ", "_")
        safe_company = "".join(c for c in job.company if c.isalnum() or c in " _-")[:20].strip().replace(" ", "_")
        file_base = f"Utkarsh_Singh_{safe_role}_{safe_company}_{run_date}"

        # Col 1: Date Found (pipeline run date)
        ws.cell(row=row, column=1, value=run_date)
        # Col 2: Posted Date (actual job posting date from scraper)
        posted = job.posted_date[:10] if job.posted_date else ""
        ws.cell(row=row, column=2, value=posted)
        # Col 3: Initial match score (true job fit, before tailoring)
        ws.cell(row=row, column=3, value=job.initial_match_score or job.match_score)
        # Col 4-7: Post-tailoring scores
        ws.cell(row=row, column=4, value=job.match_score)
        ws.cell(row=row, column=5, value=job.ats_score)
        ws.cell(row=row, column=6, value=job.hiring_manager_score)
        ws.cell(row=row, column=7, value=job.tech_recruiter_score)
        # Col 8-14: Job details
        ws.cell(row=row, column=8, value=job.title)
        ws.cell(row=row, column=9, value=job.company)
        ws.cell(row=row, column=10, value=job.location)
        ws.cell(row=row, column=11, value="Yes" if job.remote else "No")
        ws.cell(row=row, column=12, value=job.salary or "Not listed")
        ws.cell(row=row, column=13, value=job.source)
        ws.cell(row=row, column=14, value=job.matched_resume)

        # Col 15: Apply link
        apply_cell = ws.cell(row=row, column=15)
        if job.apply_url:
            apply_cell.value = "Apply"
            apply_cell.hyperlink = job.apply_url
            apply_cell.font = LINK_FONT
        else:
            apply_cell.value = "No link"

        # Col 16: Resume PDF (with S3 link if available)
        resume_cell = ws.cell(row=row, column=16)
        if job.resume_s3_url:
            resume_cell.value = f"{file_base}.pdf"
            resume_cell.hyperlink = job.resume_s3_url
            resume_cell.font = LINK_FONT
        elif job.tailored_pdf_path:
            resume_cell.value = Path(job.tailored_pdf_path).name
        else:
            resume_cell.value = "—"

        # Col 17: Cover letter (with S3 link if available)
        cl_cell = ws.cell(row=row, column=17)
        if job.cover_letter_s3_url:
            cl_cell.value = f"{file_base}_CoverLetter.pdf"
            cl_cell.hyperlink = job.cover_letter_s3_url
            cl_cell.font = LINK_FONT
        elif job.cover_letter_pdf_path:
            cl_cell.value = Path(job.cover_letter_pdf_path).name
        else:
            cl_cell.value = "—"

        # Col 18: Resume (Drive link)
        resume_drive_cell = ws.cell(row=row, column=18)
        if job.resume_drive_url:
            resume_drive_cell.value = "Open"
            resume_drive_cell.hyperlink = job.resume_drive_url
            resume_drive_cell.font = LINK_FONT
        else:
            resume_drive_cell.value = "—"

        # Col 19: Cover Letter (Drive link)
        cl_drive_cell = ws.cell(row=row, column=19)
        if job.cover_letter_drive_url:
            cl_drive_cell.value = "Open"
            cl_drive_cell.hyperlink = job.cover_letter_drive_url
            cl_drive_cell.font = LINK_FONT
        else:
            cl_drive_cell.value = "—"

        # Col 20-27: LinkedIn contacts (up to 3)
        # Contact 1: cols 20 (role), 21 (linkedin), 22 (message)
        # Contact 2: cols 23 (role), 24 (linkedin), 25 (message)
        # Contact 3: cols 26 (role), 27 (linkedin)
        for ci in range(3):
            c = contacts[ci] if ci < len(contacts) else {}
            c_role = c.get("role", "")
            c_url = c.get("search_url", "")
            c_msg = c.get("message", "")

            if ci < 2:
                # Contacts 1 & 2: role + linkedin + message (3 cols each)
                base = 20 + ci * 3  # 20 or 23
                ws.cell(row=row, column=base, value=c_role)
                li_cell = ws.cell(row=row, column=base + 1)
                if c_url:
                    li_cell.value = "Search"
                    li_cell.hyperlink = c_url
                    li_cell.font = LINK_FONT
                else:
                    li_cell.value = "—"
                ws.cell(row=row, column=base + 2, value=c_msg)
            else:
                # Contact 3: role + linkedin only (2 cols)
                ws.cell(row=row, column=26, value=c_role)
                li_cell = ws.cell(row=row, column=27)
                if c_url:
                    li_cell.value = "Search"
                    li_cell.hyperlink = c_url
                    li_cell.font = LINK_FONT
                else:
                    li_cell.value = "—"

        # Col 28-30: Application tracking
        ws.cell(row=row, column=28, value="No")   # Applied?
        ws.cell(row=row, column=29, value="")      # Applied Date
        ws.cell(row=row, column=30, value="New")   # Status

        # Col 31-32: Follow-up dates (calculated when applied date is set)
        ws.cell(row=row, column=31, value="")  # Follow-Up 1 (1 week after applied)
        ws.cell(row=row, column=32, value="")  # Follow-Up 2 (2 weeks after applied)

        # Col 33-34: Follow-up messages
        ws.cell(row=row, column=33, value=followup_1_msg)
        ws.cell(row=row, column=34, value=followup_2_msg)

        # Col 35: Apply Reminder
        reminder_cell = ws.cell(row=row, column=35)
        reminder_cell.value = "APPLY NOW!"
        reminder_cell.fill = REMINDER_FILL
        reminder_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

        # Col 36: Notes
        ws.cell(row=row, column=36, value="")

        # ── Format the row ──
        score_cols = {3, 4, 5, 6, 7}  # Match=3, Score=4, ATS=5, HM=6, TR=7
        status_col = 30
        reminder_col = 35
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.font or cell.font == Font():
                cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER
            if is_zebra and col not in score_cols and col != status_col and col != reminder_col:
                cell.fill = ZEBRA_FILL

        # Color-code scores
        _color_score_cell(ws.cell(row=row, column=3), job.initial_match_score or job.match_score)
        _color_score_cell(ws.cell(row=row, column=4), job.match_score)
        _color_score_cell(ws.cell(row=row, column=5), job.ats_score)
        _color_score_cell(ws.cell(row=row, column=6), job.hiring_manager_score)
        _color_score_cell(ws.cell(row=row, column=7), job.tech_recruiter_score)

        # Color-code status (col 30)
        ws.cell(row=row, column=30).fill = STATUS_COLORS.get("New", PatternFill())

    # Update summary
    if "Daily Summary" in wb.sheetnames:
        _update_summary(wb["Daily Summary"], new_jobs, run_date)

    # Auto-filter
    last_row = start_row + len(new_jobs) - 1
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    ws.freeze_panes = "H2"  # Freeze through Posted Date + Match + Score + ATS/HM/TR

    wb.save(str(tracker_path))
    logger.info(f"[EXCEL] Master tracker updated: {tracker_path} ({len(new_jobs)} new jobs added)")
    return str(tracker_path)


def _update_reminders(ws, run_date: str):
    """Update Apply Reminder and Follow-Up dates for existing rows.

    Detects column layout from headers to handle both old and new formats.
    """
    today = datetime.strptime(run_date, "%Y-%m-%d").date()

    # Detect column indices from headers
    col_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            col_map[header] = col

    applied_col = col_map.get("Applied?", 20)
    applied_date_col = col_map.get("Applied Date", 21)
    fu1_col = col_map.get("Follow-Up 1", 23)
    fu2_col = col_map.get("Follow-Up 2", 24)
    reminder_col = col_map.get("Apply Reminder", 27)
    date_found_col = col_map.get("Date Found", 1)

    for row in range(2, ws.max_row + 1):
        applied = ws.cell(row=row, column=applied_col).value
        applied_date_val = ws.cell(row=row, column=applied_date_col).value

        if applied and str(applied).strip().lower() == "yes" and applied_date_val:
            # Clear the reminder
            r_cell = ws.cell(row=row, column=reminder_col)
            r_cell.value = "Applied"
            r_cell.fill = PatternFill("solid", fgColor="92D050")
            r_cell.font = Font(name="Calibri", size=10, color="FFFFFF")

            # Calculate follow-up dates
            try:
                if isinstance(applied_date_val, datetime):
                    applied_date = applied_date_val.date()
                else:
                    applied_date = datetime.strptime(str(applied_date_val).strip(), "%Y-%m-%d").date()

                fu1 = applied_date + timedelta(days=7)
                fu2 = applied_date + timedelta(days=14)

                fu1_cell = ws.cell(row=row, column=fu1_col)
                fu2_cell = ws.cell(row=row, column=fu2_col)

                if not fu1_cell.value:
                    fu1_cell.value = fu1.isoformat()
                if not fu2_cell.value:
                    fu2_cell.value = fu2.isoformat()

                # Highlight follow-ups that are due
                if fu1 <= today:
                    fu1_cell.fill = FOLLOWUP_DUE
                if fu2 <= today:
                    fu2_cell.fill = FOLLOWUP_DUE

            except (ValueError, TypeError):
                pass

        elif not applied or str(applied).strip().lower() != "yes":
            # Not yet applied — keep the reminder active
            days_since = 0
            date_found = ws.cell(row=row, column=date_found_col).value
            if date_found:
                try:
                    if isinstance(date_found, datetime):
                        found = date_found.date()
                    else:
                        found = datetime.strptime(str(date_found).strip(), "%Y-%m-%d").date()
                    days_since = (today - found).days
                except (ValueError, TypeError):
                    pass

            r_cell = ws.cell(row=row, column=reminder_col)
            if days_since >= 3:
                r_cell.value = f"URGENT! ({days_since}d)"
                r_cell.fill = REMINDER_FILL
                r_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            elif days_since >= 1:
                r_cell.value = "APPLY NOW!"
                r_cell.fill = REMINDER_FILL
                r_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")


def _get_existing_keys(ws) -> set:
    """Extract title|company keys from existing rows.

    Detects column layout from header row to handle both old (Title=6, Company=7)
    and new (Title=7, Company=8) formats.
    """
    # Find Title and Company columns from header
    title_col, company_col = 7, 8  # New defaults
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == "Title":
            title_col = col
        elif header == "Company":
            company_col = col

    keys = set()
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=title_col).value or ""
        company = ws.cell(row=row, column=company_col).value or ""
        key = f"{str(title).lower().strip()}|{str(company).lower().strip()}"
        keys.add(key)
    return keys


def _color_score_cell(cell, score):
    if score >= 85:
        cell.fill = SCORE_EXCELLENT
    elif score >= 75:
        cell.fill = SCORE_GOOD
    elif score >= 60:
        cell.fill = SCORE_OK
    else:
        cell.fill = SCORE_LOW


def _setup_header(ws):
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32


def _add_data_validations(ws):
    applied_dv = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid", error="Select Yes or No",
    )
    applied_dv.sqref = "AB2:AB5000"  # Col 28 = AB (Applied?)
    ws.add_data_validation(applied_dv)

    status_dv = DataValidation(
        type="list", formula1='"New,Applied,Interview,Offer,Rejected,Withdrawn"',
        allow_blank=True, showErrorMessage=True, errorTitle="Invalid", error="Select a valid status",
    )
    status_dv.sqref = "AD2:AD5000"  # Col 30 = AD (Status)
    ws.add_data_validation(status_dv)


def _setup_summary_sheet(ws):
    headers = [
        "Date", "New Jobs", "Already Tracked", "Avg Score",
        "Avg ATS", "Avg HM", "Avg TR", "All 85+",
        "Resumes", "Cover Letters", "Top Company", "Top Role",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


def _update_summary(ws, jobs: List[Job], run_date: str):
    if not jobs:
        return

    row = ws.max_row + 1
    avg_score = sum(j.match_score for j in jobs) / len(jobs)
    avg_ats = sum(j.ats_score for j in jobs) / len(jobs)
    avg_hm = sum(j.hiring_manager_score for j in jobs) / len(jobs)
    avg_tr = sum(j.tech_recruiter_score for j in jobs) / len(jobs)
    all_85 = sum(1 for j in jobs if j.ats_score >= 85 and j.hiring_manager_score >= 85 and j.tech_recruiter_score >= 85)
    top_job = max(jobs, key=lambda j: j.match_score) if jobs else None

    ws.cell(row=row, column=1, value=run_date)
    ws.cell(row=row, column=2, value=len(jobs))
    ws.cell(row=row, column=3, value=0)
    ws.cell(row=row, column=4, value=round(avg_score, 1))
    ws.cell(row=row, column=5, value=round(avg_ats, 1))
    ws.cell(row=row, column=6, value=round(avg_hm, 1))
    ws.cell(row=row, column=7, value=round(avg_tr, 1))
    ws.cell(row=row, column=8, value=all_85)
    ws.cell(row=row, column=9, value=sum(1 for j in jobs if j.tailored_pdf_path))
    ws.cell(row=row, column=10, value=sum(1 for j in jobs if j.cover_letter_pdf_path))
    ws.cell(row=row, column=11, value=top_job.company if top_job else "N/A")
    ws.cell(row=row, column=12, value=top_job.title if top_job else "N/A")

    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGNMENT
        cell.border = THIN_BORDER
